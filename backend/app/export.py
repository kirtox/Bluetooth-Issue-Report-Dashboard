from fastapi import APIRouter, Depends, Query
from sqlalchemy.orm import Session
from .db import get_db
from .models import Report
import io
import csv
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from typing import Optional, List
from datetime import datetime
from . import crud

router = APIRouter()

# Common headers aligned with schema_report.ReportBase
HEADERS = [
    "ID", "Operator Name", "Date", "Serial Number", "OS Version", "Platform Brand", "Platform",
    "Platform Phase", "Platform BIOS", "CPU", "WLAN", "WLAN Phase",
    "BT Driver", "BT Interface", "WiFi Driver", "Audio Driver", "WRT Version", "WRT Preset",
    "MSFT Teams Version", "Scenario",
    "Mouse BT", "Mouse Brand", "Mouse", "Mouse Click Period",
    "Keyboard BT", "Keyboard Brand", "Keyboard", "Keyboard Click Period",
    "Headset BT", "Headset Brand", "Headset",
    "Speaker BT", "Speaker Brand", "Speaker",
    "Phone Brand", "Phone",
    "Device1 Brand", "Device1",
    "Modern Standby", "MS Period", "MS OS Waiting Time",
    "S4", "S4 Period", "S4 OS Waiting Time",
    "Warm Boot", "WB Period", "WB OS Waiting Time",
    "Cold Boot", "CB Period", "CB OS Waiting Time",
    "Microsoft Teams", "APM", "APM Period", "OPP", "Swift Pair",
    "Power Type", "Urgent Level", "Fix Work Week", "Fix BT Driver", "JIRA ID", "IPS ID", "HSD ID",
    "Result", "Fail Cycles", "Cycles", "Duration", "Log Path"
]

@router.get("/export/csv")
def export_csv(
    db: Session = Depends(get_db),
    search_term: Optional[str] = Query(None),
    platforms: Optional[str] = Query(None),
    results: Optional[str] = Query(None),
    wlans: Optional[str] = Query(None),
    scenarios: Optional[str] = Query(None),
    bt_drivers: Optional[str] = Query(None),
    start_date: Optional[str] = Query(None),
    end_date: Optional[str] = Query(None)
):
    platform_list = platforms.split(',') if platforms else None
    result_list = results.split(',') if results else None
    wlan_list = wlans.split(',') if wlans else None
    scenario_list = scenarios.split(',') if scenarios else None
    bt_driver_list = bt_drivers.split(',') if bt_drivers else None

    reports = crud.get_reports_filtered(
        db=db,
        search_term=search_term,
        platforms=platform_list,
        results=result_list,
        wlans=wlan_list,
        scenarios=scenario_list,
        bt_drivers=bt_driver_list,
        start_date=start_date,
        end_date=end_date,
    )

    stream = io.StringIO()
    writer = csv.writer(stream)
    writer.writerow(HEADERS)
    for r in reports:
        writer.writerow([
            r.id,
            r.op_name,
            r.date.strftime("%Y-%m-%d %H:%M:%S") if getattr(r, 'date', None) else "",
            r.serial_num,
            r.os_version,
            r.platform_brand,
            r.platform,
            r.platform_phase,
            r.platform_bios,
            r.cpu,
            r.wlan,
            r.wlan_phase,
            r.bt_driver,
            r.bt_interface,
            r.wifi_driver,
            r.audio_driver,
            r.wrt_version,
            r.wrt_preset,
            r.msft_teams_version,
            r.scenario,
            r.mouse_bt,
            r.mouse_brand,
            r.mouse,
            r.mouse_click_period,
            r.keyboard_bt,
            r.keyboard_brand,
            r.keyboard,
            r.keyboard_click_period,
            r.headset_bt,
            r.headset_brand,
            r.headset,
            r.speaker_bt,
            r.speaker_brand,
            r.speaker,
            r.phone_brand,
            r.phone,
            r.device1_brand,
            r.device1,
            r.modern_standby,
            r.ms_period,
            r.ms_os_waiting_time,
            r.s4,
            r.s4_period,
            r.s4_os_waiting_time,
            r.warm_boot,
            r.wb_period,
            r.wb_os_waiting_time,
            r.cold_boot,
            r.cb_period,
            r.cb_os_waiting_time,
            r.microsoft_teams,
            r.apm,
            r.apm_period,
            r.opp,
            r.swift_pair,
            r.power_type,
            r.urgent_level,
            r.fix_work_week,
            r.fix_bt_driver,
            r.jira_id,
            r.ips_id,
            r.hsd_id,
            r.result,
            r.fail_cycles,
            r.cycles,
            r.duration,
            r.log_path,
        ])
    stream.seek(0)
    return StreamingResponse(stream, media_type="text/csv", headers={"Content-Disposition": "attachment; filename=reports.csv"})

@router.get("/export/excel")
def export_excel(
    db: Session = Depends(get_db),
    search_term: Optional[str] = Query(None),
    platforms: Optional[str] = Query(None),
    results: Optional[str] = Query(None),
    wlans: Optional[str] = Query(None),
    scenarios: Optional[str] = Query(None),
    bt_drivers: Optional[str] = Query(None),
    start_date: Optional[str] = Query(None),
    end_date: Optional[str] = Query(None)
):
    platform_list = platforms.split(',') if platforms else None
    result_list = results.split(',') if results else None
    wlan_list = wlans.split(',') if wlans else None
    scenario_list = scenarios.split(',') if scenarios else None
    bt_driver_list = bt_drivers.split(',') if bt_drivers else None

    reports = crud.get_reports_filtered(
        db=db,
        search_term=search_term,
        platforms=platform_list,
        results=result_list,
        wlans=wlan_list,
        scenarios=scenario_list,
        bt_drivers=bt_driver_list,
        start_date=start_date,
        end_date=end_date,
    )

    # Build Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Reports"
    
    # Defined headers (aligned with schema_report.ReportBase)
    headers = HEADERS
    
    # Set styles of headers
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Write data
    for row, report in enumerate(reports, 2):
        ws.cell(row=row, column=1, value=report.id)
        ws.cell(row=row, column=2, value=report.op_name)
        ws.cell(row=row, column=3, value=report.date.strftime("%Y-%m-%d %H:%M:%S") if getattr(report, 'date', None) else "")
        ws.cell(row=row, column=4, value=report.serial_num)
        ws.cell(row=row, column=5, value=report.os_version)
        ws.cell(row=row, column=6, value=report.platform_brand)
        ws.cell(row=row, column=7, value=report.platform)
        ws.cell(row=row, column=8, value=report.platform_phase)
        ws.cell(row=row, column=9, value=report.platform_bios)
        ws.cell(row=row, column=10, value=report.cpu)
        ws.cell(row=row, column=11, value=report.wlan)
        ws.cell(row=row, column=12, value=report.wlan_phase)
        ws.cell(row=row, column=13, value=report.bt_driver)
        ws.cell(row=row, column=14, value=report.bt_interface)
        ws.cell(row=row, column=15, value=report.wifi_driver)
        ws.cell(row=row, column=16, value=report.audio_driver)
        ws.cell(row=row, column=17, value=report.wrt_version)
        ws.cell(row=row, column=18, value=report.wrt_preset)
        ws.cell(row=row, column=19, value=report.msft_teams_version)
        ws.cell(row=row, column=20, value=report.scenario)
        ws.cell(row=row, column=21, value=report.mouse_bt)
        ws.cell(row=row, column=22, value=report.mouse_brand)
        ws.cell(row=row, column=23, value=report.mouse)
        ws.cell(row=row, column=24, value=report.mouse_click_period)
        ws.cell(row=row, column=25, value=report.keyboard_bt)
        ws.cell(row=row, column=26, value=report.keyboard_brand)
        ws.cell(row=row, column=27, value=report.keyboard)
        ws.cell(row=row, column=28, value=report.keyboard_click_period)
        ws.cell(row=row, column=29, value=report.headset_bt)
        ws.cell(row=row, column=30, value=report.headset_brand)
        ws.cell(row=row, column=31, value=report.headset)
        ws.cell(row=row, column=32, value=report.speaker_bt)
        ws.cell(row=row, column=33, value=report.speaker_brand)
        ws.cell(row=row, column=34, value=report.speaker)
        ws.cell(row=row, column=35, value=report.phone_brand)
        ws.cell(row=row, column=36, value=report.phone)
        ws.cell(row=row, column=37, value=report.device1_brand)
        ws.cell(row=row, column=38, value=report.device1)
        ws.cell(row=row, column=39, value=report.modern_standby)
        ws.cell(row=row, column=40, value=report.ms_period)
        ws.cell(row=row, column=41, value=report.ms_os_waiting_time)
        ws.cell(row=row, column=42, value=report.s4)
        ws.cell(row=row, column=43, value=report.s4_period)
        ws.cell(row=row, column=44, value=report.s4_os_waiting_time)
        ws.cell(row=row, column=45, value=report.warm_boot)
        ws.cell(row=row, column=46, value=report.wb_period)
        ws.cell(row=row, column=47, value=report.wb_os_waiting_time)
        ws.cell(row=row, column=48, value=report.cold_boot)
        ws.cell(row=row, column=49, value=report.cb_period)
        ws.cell(row=row, column=50, value=report.cb_os_waiting_time)
        ws.cell(row=row, column=51, value=report.microsoft_teams)
        ws.cell(row=row, column=52, value=report.apm)
        ws.cell(row=row, column=53, value=report.apm_period)
        ws.cell(row=row, column=54, value=report.opp)
        ws.cell(row=row, column=55, value=report.swift_pair)
        ws.cell(row=row, column=56, value=report.power_type)
        ws.cell(row=row, column=57, value=report.urgent_level)
        ws.cell(row=row, column=58, value=report.fix_work_week)
        ws.cell(row=row, column=59, value=report.fix_bt_driver)
        ws.cell(row=row, column=60, value=report.jira_id)
        ws.cell(row=row, column=61, value=report.ips_id)
        ws.cell(row=row, column=62, value=report.hsd_id)
        ws.cell(row=row, column=63, value=report.result)
        ws.cell(row=row, column=64, value=report.fail_cycles)
        ws.cell(row=row, column=65, value=report.cycles)
        ws.cell(row=row, column=66, value=report.duration)
        ws.cell(row=row, column=67, value=report.log_path)
    
    # Auto adjust the width of columns 
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50) # Maximum width: 50
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to memory
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Generate the filename
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"reports_{timestamp}.xlsx"
    
    return StreamingResponse(
        output, 
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )